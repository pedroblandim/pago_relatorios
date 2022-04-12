# -*- coding: utf-8 -*-
from openpyxl import Workbook
from datetime import date
from openpyxl.styles import Font, PatternFill

import os

class PaymentSheet:
    def __init__(self, request) -> None:
        data = request.json
        self.payments = data["Pagamentos"]
        self.company_name = data["Empresa"]

    def create_sheet(self, sheet_path, sheet_file_name):
        workbook = Workbook()
        planilha = self.__set_format_xlsx(workbook)

        self.__add_values_xlsx(planilha)
        self.__soma_total(planilha)

        workbook.save(os.path.join(sheet_path, sheet_file_name))

    def __set_format_xlsx(self, workbook):
        data_atual = date.today()
        font_style_title = Font(bold=True)
        font_style = Font(bold=True, color="FFFFFF")
        fill_pattern = PatternFill(patternType='solid', fgColor="45818D")

        planilha = workbook.worksheets[0]

        planilha.title = 'Relatório de Pagamentos'

        planilha['a1'] = self.company_name.upper()
        planilha['b1'] = data_atual.strftime('%d/%m/%Y')
        planilha['a1'].font = planilha['b1'].font = font_style_title

        planilha['a3'] = 'NOME'
        planilha['b3'] = 'VALOR'
        planilha['c3'] = 'TIPO'
        planilha['a3'].font = planilha['b3'].font = planilha['c3'].font = font_style
        planilha['a3'].fill = planilha['b3'].fill =planilha['c3'].fill = fill_pattern

        return planilha

    def __add_values_xlsx(self, planilha):
        starting_row = 4
        for item in self.payments:
            planilha.cell(row=starting_row, column=1).value = item['name']
            planilha.cell(row=starting_row, column=2).value = item['value']
            planilha.cell(row=starting_row, column=3).value = item['type']
            starting_row+=1
        return planilha

    def __soma_total(self, planilha):
        soma_valor = 0
        starting_row = 4
        for item in self.payments:
            soma_valor += item['value']
            starting_row += 1
        planilha.cell(row = starting_row + 1, column=1).value = 'Total'
        planilha.cell(row = starting_row + 1, column=2).value = soma_valor
        return planilha
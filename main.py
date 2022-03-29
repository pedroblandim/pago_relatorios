# -*- coding: utf-8 -*-
from openpyxl import Workbook
from datetime import date
from openpyxl.styles import Font, PatternFill

import os
import pathlib

from flask import Flask, send_from_directory, request

app = Flask(__name__)

SHEET_PATH = pathlib.Path().resolve()
SHEET_NAME = "Relatório de Pagamentos.xlsx"

@app.post("/")
def hello_world():
    data = request.json
    
    payments = data["Pagamentos"]
    company_name = data["Empresa"]

    create_sheet(company_name, payments)
    return send_from_directory(SHEET_PATH, SHEET_NAME, as_attachment=True)

def create_xlsx():
    return Workbook()

def set_format_xlsx(workbook, nome_empresa):
    data_atual = date.today()
    font_style_title = Font(bold=True)
    font_style = Font(bold=True, color="FFFFFF")
    fill_pattern = PatternFill(patternType='solid', fgColor="45818D")

    planilha = workbook.worksheets[0]

    planilha.title = 'Relatório de Pagamentos'

    planilha['a1'] = nome_empresa.upper()
    planilha['b1'] = data_atual.strftime('%d/%m/%Y')
    planilha['a1'].font = planilha['b1'].font = font_style_title

    planilha['a3'] = 'NOME'
    planilha['b3'] = 'VALOR'
    planilha['c3'] = 'TIPO'
    planilha['a3'].font = planilha['b3'].font = planilha['c3'].font = font_style
    planilha['a3'].fill = planilha['b3'].fill =planilha['c3'].fill = fill_pattern

    return planilha

def add_values_xlsx(planilha, payments):
    starting_row = 4
    for item in payments:
        planilha.cell(row=starting_row, column=1).value = item['name']
        planilha.cell(row=starting_row, column=2).value = item['value']
        planilha.cell(row=starting_row, column=3).value = item['type']
        starting_row+=1
    return planilha

def soma_total(planilha, payments):
    soma_valor = 0
    starting_row = 4
    for item in payments:
        soma_valor += item['value']
        starting_row += 1
    planilha.cell(row = starting_row + 1, column=1).value = 'Total'
    planilha.cell(row = starting_row + 1, column=2).value = soma_valor
    return planilha

def create_sheet(company_name, payments):
    workbook = create_xlsx()
    planilha = set_format_xlsx(workbook, company_name)

    add_values_xlsx(planilha, payments)
    soma_total(planilha, payments)

    workbook.save(os.path.join(SHEET_PATH, SHEET_NAME))